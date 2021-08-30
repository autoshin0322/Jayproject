import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active


ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["A2"] = "=Sum(1,2,3)" # 1 + 2 + 3 = 6
ws["A3"] = "=AVERAGE(1,2,3)" # 2 (평균)

ws["A4"] = 10
ws["a5"] = 20
ws["a6"] = "=Sum(a4:a5)"

wb.save("sample_formular.xlsx")